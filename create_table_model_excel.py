"""
创建数据表模型Excel文件
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_table_model_excel():
    """创建数据表模型Excel文件"""
    
    # 定义所有表的结构
    tables = [
        {
            'name': 'user',
            'description': '用户表',
            'columns': [
                {'name': 'user_id', 'type': 'VARCHAR(32)', 'pk': True, 'fk': False, 'nullable': False, 'default': '', 'comment': '用户ID，全局唯一'},
                {'name': 'user_name', 'type': 'VARCHAR(50)', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '用户姓名'},
                {'name': 'user_type', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': 'PERSONAL', 'comment': '用户类型：PERSONAL/INSTITUTION'},
                {'name': 'user_status', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': 'ACTIVE', 'comment': '用户状态：ACTIVE/INACTIVE/FROZEN'},
                {'name': 'identity_no', 'type': 'VARCHAR(30)', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '身份证/机构代码'},
                {'name': 'phone', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '手机号'},
                {'name': 'email', 'type': 'VARCHAR(100)', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '邮箱'},
                {'name': 'create_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': False, 'default': 'CURRENT_TIMESTAMP', 'comment': '创建时间'},
                {'name': 'update_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': False, 'default': 'CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP', 'comment': '更新时间'},
            ]
        },
        {
            'name': 'user_bank_card',
            'description': '用户银行卡表',
            'columns': [
                {'name': 'card_id', 'type': 'VARCHAR(32)', 'pk': True, 'fk': False, 'nullable': False, 'default': '', 'comment': '银行卡ID'},
                {'name': 'user_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '用户ID，关联user.user_id'},
                {'name': 'bank_code', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '银行代码'},
                {'name': 'bank_name', 'type': 'VARCHAR(50)', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '银行名称'},
                {'name': 'card_no', 'type': 'VARCHAR(30)', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '银行卡号（加密存储）'},
                {'name': 'card_type', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': 'DEBIT', 'comment': '卡类型：DEBIT/CREDIT'},
                {'name': 'card_status', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': 'ACTIVE', 'comment': '状态：ACTIVE/INACTIVE/FROZEN'},
                {'name': 'bind_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': False, 'default': 'CURRENT_TIMESTAMP', 'comment': '绑定时间'},
                {'name': 'is_default', 'type': 'BOOLEAN', 'pk': False, 'fk': False, 'nullable': False, 'default': 'FALSE', 'comment': '是否默认卡'},
            ]
        },
        {
            'name': 'user_balance',
            'description': '用户资金余额表',
            'columns': [
                {'name': 'balance_id', 'type': 'VARCHAR(32)', 'pk': True, 'fk': False, 'nullable': False, 'default': '', 'comment': '余额ID'},
                {'name': 'user_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '用户ID，关联user.user_id'},
                {'name': 'available_balance', 'type': 'DECIMAL(18,4)', 'pk': False, 'fk': False, 'nullable': False, 'default': '0', 'comment': '可用余额'},
                {'name': 'frozen_balance', 'type': 'DECIMAL(18,4)', 'pk': False, 'fk': False, 'nullable': False, 'default': '0', 'comment': '冻结余额'},
                {'name': 'total_balance', 'type': 'DECIMAL(18,4)', 'pk': False, 'fk': False, 'nullable': False, 'default': '0', 'comment': '总余额（计算字段）'},
                {'name': 'last_update', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': False, 'default': 'CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP', 'comment': '最后更新时间'},
            ]
        },
        {
            'name': 'capital_change_entrust',
            'description': '资金变动委托表',
            'columns': [
                {'name': 'entrust_id', 'type': 'VARCHAR(32)', 'pk': True, 'fk': False, 'nullable': False, 'default': '', 'comment': '委托ID'},
                {'name': 'user_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '用户ID'},
                {'name': 'card_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '银行卡ID'},
                {'name': 'change_type', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '变动类型：IN/OUT'},
                {'name': 'amount', 'type': 'DECIMAL(18,4)', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '变动金额'},
                {'name': 'third_party_no', 'type': 'VARCHAR(50)', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '第三方流水号'},
                {'name': 'status', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': 'PENDING', 'comment': '状态：PENDING/PROCESSING/SUCCESS/FAILED'},
                {'name': 'request_data', 'type': 'JSON', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '请求数据'},
                {'name': 'response_data', 'type': 'JSON', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '响应数据'},
                {'name': 'create_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': False, 'default': 'CURRENT_TIMESTAMP', 'comment': '创建时间'},
                {'name': 'process_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '处理时间'},
                {'name': 'complete_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '完成时间'},
                {'name': 'error_msg', 'type': 'TEXT', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '错误信息'},
            ]
        },
        {
            'name': 'capital_settlement',
            'description': '资金清算表',
            'columns': [
                {'name': 'settlement_id', 'type': 'VARCHAR(32)', 'pk': True, 'fk': False, 'nullable': False, 'default': '', 'comment': '清算ID'},
                {'name': 'entrust_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '委托ID，关联capital_change_entrust.entrust_id'},
                {'name': 'user_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '用户ID'},
                {'name': 'balance_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '余额ID'},
                {'name': 'settlement_type', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '清算类型：CAPITAL_IN/CAPITAL_OUT'},
                {'name': 'amount', 'type': 'DECIMAL(18,4)', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '清算金额'},
                {'name': 'result_status', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '结果状态：SUCCESS/FAILED'},
                {'name': 'settlement_data', 'type': 'JSON', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '清算数据'},
                {'name': 'settlement_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': False, 'default': 'CURRENT_TIMESTAMP', 'comment': '清算时间'},
                {'name': 'remark', 'type': 'VARCHAR(500)', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '备注'},
            ]
        },
        {
            'name': 'fund_account',
            'description': '基金账户表',
            'columns': [
                {'name': 'fund_account_id', 'type': 'VARCHAR(32)', 'pk': True, 'fk': False, 'nullable': False, 'default': '', 'comment': '基金账户ID'},
                {'name': 'user_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '用户ID，关联user.user_id'},
                {'name': 'account_no', 'type': 'VARCHAR(30)', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '基金账户号'},
                {'name': 'account_type', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': 'INDIVIDUAL', 'comment': '账户类型：INDIVIDUAL/INSTITUTION'},
                {'name': 'account_status', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': 'ACTIVE', 'comment': '状态：ACTIVE/INACTIVE/FROZEN'},
                {'name': 'open_date', 'type': 'DATE', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '开户日期'},
                {'name': 'close_date', 'type': 'DATE', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '销户日期'},
                {'name': 'create_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': False, 'default': 'CURRENT_TIMESTAMP', 'comment': '创建时间'},
            ]
        },
        {
            'name': 'fund_account_entrust',
            'description': '基金账户委托表',
            'columns': [
                {'name': 'entrust_id', 'type': 'VARCHAR(32)', 'pk': True, 'fk': False, 'nullable': False, 'default': '', 'comment': '委托ID'},
                {'name': 'user_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '用户ID'},
                {'name': 'account_type', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': 'INDIVIDUAL', 'comment': '账户类型'},
                {'name': 'source_channel', 'type': 'VARCHAR(30)', 'pk': False, 'fk': False, 'nullable': False, 'default': 'WEB', 'comment': '来源渠道：WEB/APP/BANK'},
                {'name': 'status', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': 'PENDING', 'comment': '状态：PENDING/PROCESSING/SUCCESS/FAILED'},
                {'name': 'request_data', 'type': 'JSON', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '请求数据'},
                {'name': 'response_data', 'type': 'JSON', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '响应数据'},
                {'name': 'create_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': False, 'default': 'CURRENT_TIMESTAMP', 'comment': '创建时间'},
                {'name': 'process_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '处理时间'},
                {'name': 'complete_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '完成时间'},
                {'name': 'error_msg', 'type': 'TEXT', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '错误信息'},
            ]
        },
        {
            'name': 'fund_product',
            'description': '基金产品表',
            'columns': [
                {'name': 'product_id', 'type': 'VARCHAR(32)', 'pk': True, 'fk': False, 'nullable': False, 'default': '', 'comment': '产品ID'},
                {'name': 'product_code', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '产品代码'},
                {'name': 'product_name', 'type': 'VARCHAR(100)', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '产品名称'},
                {'name': 'product_type', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': 'EQUITY', 'comment': '产品类型：EQUITY/BOND/MIXED/MONETARY'},
                {'name': 'product_status', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': 'ACTIVE', 'comment': '状态：ACTIVE/INACTIVE'},
                {'name': 'risk_level', 'type': 'VARCHAR(10)', 'pk': False, 'fk': False, 'nullable': False, 'default': 'R3', 'comment': '风险等级：R1-R5'},
                {'name': 'fund_company', 'type': 'VARCHAR(50)', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '基金公司'},
                {'name': 'issue_date', 'type': 'DATE', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '发行日期'},
                {'name': 'create_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': False, 'default': 'CURRENT_TIMESTAMP', 'comment': '创建时间'},
            ]
        },
        {
            'name': 'fund_net_value',
            'description': '基金单位净值表',
            'columns': [
                {'name': 'nav_id', 'type': 'VARCHAR(32)', 'pk': True, 'fk': False, 'nullable': False, 'default': '', 'comment': '净值ID'},
                {'name': 'product_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '产品ID，关联fund_product.product_id'},
                {'name': 'net_value', 'type': 'DECIMAL(10,4)', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '单位净值'},
                {'name': 'accumulated_nav', 'type': 'DECIMAL(10,4)', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '累计净值'},
                {'name': 'nav_date', 'type': 'DATE', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '净值日期'},
                {'name': 'create_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': False, 'default': 'CURRENT_TIMESTAMP', 'comment': '创建时间'},
            ]
        },
        {
            'name': 'fund_transaction_entrust',
            'description': '基金交易委托表',
            'columns': [
                {'name': 'entrust_id', 'type': 'VARCHAR(32)', 'pk': True, 'fk': False, 'nullable': False, 'default': '', 'comment': '委托ID'},
                {'name': 'user_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '用户ID'},
                {'name': 'fund_account_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '基金账户ID'},
                {'name': 'product_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '产品ID'},
                {'name': 'transaction_type', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '交易类型：SUBSCRIBE/REDEEM'},
                {'name': 'amount', 'type': 'DECIMAL(18,4)', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '交易金额（申购用）'},
                {'name': 'share', 'type': 'DECIMAL(18,4)', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '交易份额（赎回用）'},
                {'name': 'nav', 'type': 'DECIMAL(10,4)', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '交易净值'},
                {'name': 'fee', 'type': 'DECIMAL(18,4)', 'pk': False, 'fk': False, 'nullable': False, 'default': '0', 'comment': '手续费'},
                {'name': 'status', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': 'PENDING', 'comment': '状态：PENDING/PROCESSING/SUCCESS/FAILED'},
                {'name': 'request_data', 'type': 'JSON', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '请求数据'},
                {'name': 'response_data', 'type': 'JSON', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '响应数据'},
                {'name': 'create_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': False, 'default': 'CURRENT_TIMESTAMP', 'comment': '创建时间'},
                {'name': 'process_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '处理时间'},
                {'name': 'complete_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '完成时间'},
                {'name': 'error_msg', 'type': 'TEXT', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '错误信息'},
            ]
        },
        {
            'name': 'fund_transaction_confirm',
            'description': '交易确认表',
            'columns': [
                {'name': 'confirm_id', 'type': 'VARCHAR(32)', 'pk': True, 'fk': False, 'nullable': False, 'default': '', 'comment': '确认ID'},
                {'name': 'entrust_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '委托ID，关联fund_transaction_entrust.entrust_id'},
                {'name': 'user_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '用户ID'},
                {'name': 'fund_account_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '基金账户ID'},
                {'name': 'product_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '产品ID'},
                {'name': 'confirm_type', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '确认类型：SUBSCRIBE/REDEEM'},
                {'name': 'result_status', 'type': 'VARCHAR(20)', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '结果状态：SUCCESS/FAILED'},
                {'name': 'confirm_data', 'type': 'JSON', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '确认数据'},
                {'name': 'confirm_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': False, 'default': 'CURRENT_TIMESTAMP', 'comment': '确认时间'},
                {'name': 'remark', 'type': 'VARCHAR(500)', 'pk': False, 'fk': False, 'nullable': True, 'default': '', 'comment': '备注'},
            ]
        },
        {
            'name': 'fund_share',
            'description': '基金份额表',
            'columns': [
                {'name': 'share_id', 'type': 'VARCHAR(32)', 'pk': True, 'fk': False, 'nullable': False, 'default': '', 'comment': '份额ID'},
                {'name': 'fund_account_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '基金账户ID'},
                {'name': 'product_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '产品ID'},
                {'name': 'total_share', 'type': 'DECIMAL(18,4)', 'pk': False, 'fk': False, 'nullable': False, 'default': '0', 'comment': '总份额'},
                {'name': 'available_share', 'type': 'DECIMAL(18,4)', 'pk': False, 'fk': False, 'nullable': False, 'default': '0', 'comment': '可用份额'},
                {'name': 'frozen_share', 'type': 'DECIMAL(18,4)', 'pk': False, 'fk': False, 'nullable': False, 'default': '0', 'comment': '冻结份额'},
                {'name': 'last_update', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': False, 'default': 'CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP', 'comment': '最后更新时间'},
            ]
        },
        {
            'name': 'user_total_asset',
            'description': '用户总资产表',
            'columns': [
                {'name': 'asset_id', 'type': 'VARCHAR(32)', 'pk': True, 'fk': False, 'nullable': False, 'default': '', 'comment': '资产ID'},
                {'name': 'user_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '用户ID'},
                {'name': 'total_asset', 'type': 'DECIMAL(18,4)', 'pk': False, 'fk': False, 'nullable': False, 'default': '0', 'comment': '总资产'},
                {'name': 'total_fund_asset', 'type': 'DECIMAL(18,4)', 'pk': False, 'fk': False, 'nullable': False, 'default': '0', 'comment': '基金总资产'},
                {'name': 'total_balance', 'type': 'DECIMAL(18,4)', 'pk': False, 'fk': False, 'nullable': False, 'default': '0', 'comment': '总余额'},
                {'name': 'calc_date', 'type': 'DATE', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '计算日期'},
                {'name': 'create_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': False, 'default': 'CURRENT_TIMESTAMP', 'comment': '创建时间'},
            ]
        },
        {
            'name': 'user_fund_asset',
            'description': '用户基金资产表',
            'columns': [
                {'name': 'fund_asset_id', 'type': 'VARCHAR(32)', 'pk': True, 'fk': False, 'nullable': False, 'default': '', 'comment': '基金资产ID'},
                {'name': 'user_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '用户ID'},
                {'name': 'product_id', 'type': 'VARCHAR(32)', 'pk': False, 'fk': True, 'nullable': False, 'default': '', 'comment': '产品ID'},
                {'name': 'fund_share', 'type': 'DECIMAL(18,4)', 'pk': False, 'fk': False, 'nullable': False, 'default': '0', 'comment': '持有份额'},
                {'name': 'fund_value', 'type': 'DECIMAL(18,4)', 'pk': False, 'fk': False, 'nullable': False, 'default': '0', 'comment': '基金价值'},
                {'name': 'nav', 'type': 'DECIMAL(10,4)', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '计算净值'},
                {'name': 'calc_date', 'type': 'DATE', 'pk': False, 'fk': False, 'nullable': False, 'default': '', 'comment': '计算日期'},
                {'name': 'create_time', 'type': 'DATETIME', 'pk': False, 'fk': False, 'nullable': False, 'default': 'CURRENT_TIMESTAMP', 'comment': '创建时间'},
            ]
        },
    ]
    
    # 创建Workbook
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认sheet
    
    # 样式定义
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    title_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 为每个表创建一个sheet
    for table in tables:
        ws = wb.create_sheet(title=table['name'])
        
        # 表标题
        ws.merge_cells('A1:H1')
        ws['A1'] = f"{table['name']} - {table['description']}"
        ws['A1'].fill = title_fill
        ws['A1'].font = title_font
        ws['A1'].alignment = center_alignment
        ws.row_dimensions[1].height = 25
        
        # 表头
        headers = ['字段名', '数据类型', '主键', '外键', '允许空', '默认值', '说明']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        ws.row_dimensions[2].height = 20
        
        # 数据行
        for row_idx, col in enumerate(table['columns'], 3):
            col_data = [
                col['name'],
                col['type'],
                '✓' if col['pk'] else '',
                '✓' if col['fk'] else '',
                '✓' if col['nullable'] else '',
                col['default'],
                col['comment']
            ]
            
            for col_idx, value in enumerate(col_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                if col_idx in [3, 4, 5]:  # 主键、外键、允许空列居中
                    cell.alignment = center_alignment
        
        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 50
    
    # 保存文件
    filename = 'database/表模型.xlsx'
    wb.save(filename)
    print(f"✓ Excel文件已创建: {filename}")
    print(f"  共 {len(tables)} 个数据表")


if __name__ == "__main__":
    create_table_model_excel()


